# -*- coding: utf-8 -*-
"""
Created on Thu Feb  9 11:09:40 2023

@author: YOGB
"""
import streamlit as st
import pandas as pd
import openpyxl
import numpy as np
from streamlit_extras.app_logo import add_logo
import utm
import folium
from streamlit_folium import st_folium

st.set_page_config(page_title="wibopargen"
                    ,layout="wide"
                   )

add_logo("https://climateapp.nl/img/about/logo-witteveenbos.png")

col1, col2 = st.columns([1,3])

with col1:
    st.title("Welcome!")
    st.markdown("""
                This application is intended for easy visualization of geotechnical parameters, either from 
                lab test results or from correlations. Creation of bar chart for grain size distribution, scatter plots, 
                and histogram chart aims to make a better engineering judgment in determining the final parameter set.
                
                """)
    st.warning("Make sure your WIBOGINA input file already contains the lab test result under sheet 'labtest'.", icon="⚠️")
    form = st.form(key='strati')
    file = form.file_uploader("Upload the WIBOGINA input spreadsheet here:", type=["xlsx"])
    utm_zone = form.number_input("UTM zone:", min_value=46, max_value=54)
    utm_hemi = form.selectbox("UTM hemisphere:", options=["North","South"])
    submit = form.form_submit_button('Upload')
    
    col_lab = "A:AX"
    col_cor = "A:K,AY:BZ"
    
if submit:
    dflab = pd.read_excel(file, sheet_name="labtest", usecols=col_lab, header=None)
    dflab.columns = dflab.loc[0:1].apply(' '.join)
    dflab = dflab.loc[2:].reset_index(drop=True)
    dfcor = pd.read_excel(file, sheet_name="labtest", usecols=col_cor, header=None)
    dfcor.columns = dfcor.loc[0:1].apply(' '.join)
    dfcor = dfcor.loc[2:].reset_index(drop=True)
    wb = openpyxl.load_workbook(file, data_only=True)
    dfstrati = pd.read_excel(file, sheet_name="stratigraphy")
    st.session_state['wb'] = wb
    st.session_state['utmz'] = utm_zone
    st.session_state['utmh'] = utm_hemi
    st.session_state['dflab'] = dflab
    st.session_state['dfcor'] = dfcor
    st.session_state["dfstrati"] = dfstrati

with col2:
    if "wb" in st.session_state:
        st.write("# Borehole/CPT Location")
    
        # SESSION STATE ===============================================================
        wb = st.session_state['wb']
        utm_zone = st.session_state['utmz'] 
        utm_hemi = st.session_state['utmh']
    
        # READING INPUT SPREADSHEET ===================================================
        sheet_bhmgr = wb['borehole_manager']
        sheet_cptmgr = wb['cpt_manager']
    
        BH_n = sheet_bhmgr.max_row - 1
        BH_ID,BH_X,BH_Y,BH_Z,BH_GWL,BH_cat = [],[],[],[],[],[]
        for i in range(BH_n):
            BH_ID.append(sheet_bhmgr.cell(i+2,2).value)
            BH_X.append(sheet_bhmgr.cell(i+2,3).value)
            BH_Y.append(sheet_bhmgr.cell(i+2,4).value)
            BH_Z.append(sheet_bhmgr.cell(i+2,5).value)
            BH_GWL.append(sheet_bhmgr.cell(i+2,6).value)
            if BH_ID[i] != None:
              BH_cat.append("BH")
    
        BH_X = np.array(BH_X)
        BH_Y = np.array(BH_Y)
        BH_Z = np.array(BH_Z)
        BH_GWL = np.array(BH_GWL)
        BH_ID = np.array(BH_ID)
        BH_cat = np.array(BH_cat)
    
        CPT_n = sheet_cptmgr.max_row - 1
        CPT_ID, CPT_X, CPT_Y, CPT_Z, CPT_GWL, CPT_cat = [],[],[],[],[],[]
        for i in range(CPT_n):
            CPT_ID.append(sheet_cptmgr.cell(i+2,2).value)
            CPT_X.append(sheet_cptmgr.cell(i+2,3).value)
            CPT_Y.append(sheet_cptmgr.cell(i+2,4).value)
            CPT_Z.append(sheet_cptmgr.cell(i+2,5).value)
            CPT_GWL.append(sheet_cptmgr.cell(i+2,6).value) 
            if CPT_ID[i] != None:
              CPT_cat.append("CPT")
    
        CPT_X = np.array(CPT_X)
        CPT_Y = np.array(CPT_Y)
        CPT_Z = np.array(CPT_Z)
        CPT_GWL = np.array(CPT_GWL)
        CPT_ID = np.array(CPT_ID)
        CPT_cat = np.array(CPT_cat)
    
        ALL_X = np.concatenate((BH_X, CPT_X))
        ALL_Y = np.concatenate((BH_Y, CPT_Y))
        ALL_Z = np.concatenate((BH_Z, CPT_Z))
        ALL_GWL = np.concatenate((BH_GWL, CPT_GWL))
        ALL_ID = np.concatenate((BH_ID, CPT_ID))
        ALL_cat = np.concatenate((BH_cat, CPT_cat))
    
        ALL_X = [i for i in ALL_X if i != None]
        ALL_Y = [i for i in ALL_Y if i != None]
        ALL_Z = [i for i in ALL_Z if i != None]
        ALL_GWL = [i for i in ALL_GWL if i != None]
        ALL_ID = [i for i in ALL_ID if i != None]
        ALL_cat = [i for i in ALL_cat if i != None]
    
        data = {'ID' : ALL_ID,
                'X_UTM' : ALL_X,
                'Y_UTM' : ALL_Y,
                'Z' : ALL_Z,
                'GWL' : ALL_GWL,
                'CLASS' : ALL_cat}
    
        df = pd.DataFrame(data)
    
        latlon, lat, lon = [],[],[]
        for i in range(len(ALL_X)):
            #print(df['ID'].iloc[i])
            if utm_hemi == 'North':
                hemis = True
            elif utm_hemi == 'South':
                hemis = False
            latlon.append(utm.to_latlon(ALL_X[i], ALL_Y[i], utm_zone, northern=hemis))
            lat.append(latlon[i][0])
            lon.append(latlon[i][1])
    
        df['lat']=lat
        df['lon']=lon
    
        latmean = df['lat'].mean()
        lonmean = df['lon'].mean()
        location = [latmean, lonmean]
        m = folium.Map(location=location, zoom_start=15
                        # , tiles="CartoDB positron"
                       )
    
        icons = []
        for i in range(0,len(df)):
            if df['CLASS'].iloc[i] == "BH":
                icons.append(folium.Icon(color='black'))
            elif df['CLASS'].iloc[i] == "CPT":
                icons.append(folium.Icon(color='red'))
    
        for i in range(0,len(df)):
            folium.Marker([df['lat'].iloc[i], df['lon'].iloc[i]], popup=df['ID'].iloc[i], icon=icons[i]).add_to(m)
    
        output = st_folium(m, width=1200, height=800, returned_objects=[])
    
    elif "wb" not in st.session_state:
        pass
